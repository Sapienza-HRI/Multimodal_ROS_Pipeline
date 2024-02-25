#!/usr/bin/env python3

import os
import torch
import autogluon
import openpyxl
import stanza

from openpyxl import load_workbook
from autogluon.text import TextPredictor
from autogluon.vision import ImagePredictor, ImageDataset
from deepmultilingualpunctuation import PunctuationModel

from dataclasses import dataclass
from torchtext.vocab import vocab, Vectors
#from transformers import BertTokenizer, BertModel, BertConfig, pipeline
from transformers import AutoModel, AutoTokenizer, AutoConfig

import rospy
from datetime import datetime
from natural_language_understanding.msg import Utterance
#from canopies_utils.srl_file import *
from canopies_utils.srl_file_xlm_roberta import *
from decision_manager.msg import FrameGroupMessage
from dialog_manager.msg import FrameMessage

#torch.cuda.empty_cache()
os.environ['CUDA_VISIBLE_DEVICES'] = '-1'


# Constants
DEVICE = "cpu"
BATCH_SIZE = 2


# TODO: transform get_user_id() function into a ROS service call between stt and nlu

def get_user_id():

    '''

    The user is asked to provide an ID (it must be the same provided in the stt_file)

    Args: -

    Returns:
        personID (float): The ID provided by the user

    '''

    # Provide an ID to the user
    print("Provide the ID assigned to the user")
    ID = float(input())
    personID = str(ID)

    return personID



def translate_sentence(text, language):

    '''

    The sentence is translated in case of a language different from english to allow SRL to correctly 
    predict the frame and arguments

    Args: 
        text (str): Sentence in the original language
        language (str): Language in which the user is interacting

    Returns:
        sentence (str): Sentence translated in english

    '''

    if language == "french":
        model_checkpoint = "Helsinki-NLP/opus-mt-fr-it" #"Helsinki-NLP/opus-mt-fr-en"  
    elif language == "german":
        model_checkpoint = "Helsinki-NLP/opus-mt-de-it" # "Helsinki-NLP/opus-mt-de-en" 
    elif language == "spanish":
        model_checkpoint = "Helsinki-NLP/opus-mt-es-it" #"Helsinki-NLP/opus-mt-es-en" 
        

    translator = pipeline("translation", model=model_checkpoint)
    result = translator(text.strip())
    out_sentence = result[0]['translation_text']
    #print(out_sentence)
    out_language = "italian" #"english"
    
    return out_sentence, out_language


'''

SRL SENTENCE PREDICTION

'''

class SRLPrediction():

    def __init__(self, device: str):
        self.device: str = device
        self.nlp = stanza.Pipeline(lang='it', processors='tokenize,mwt,pos,lemma,depparse') #en

        # Load the SRL dictionary
        store_dataset_train = torch.load("srl_model/Italian/dict_vocabs.pth", map_location=self.device)
        
        self.vocab_words = store_dataset_train["vocab_words"]
        self.vocab_pos_tags = store_dataset_train["vocab_pos_tags"]
        self.vocab_lemmas = store_dataset_train["vocab_lemmas"]
        self.vocab_predicates = store_dataset_train["vocab_predicates"]
        self.vocab_label = store_dataset_train["vocab_label"]
        net_configuration = net_configurator(use_bert_embeddings=USE_BERT_EMBEDDINGS, use_crf=USE_CRF, use_biaffine_layer=USE_BIAFFINE_LAYER, use_pretrained=False, use_dependecy_heads=USE_DEPENDENCY_HEADS, use_predicates=False, use_syntagnet=USE_SYNTAGNET)

        # -- BERT --
        self.model_name: str = 'xlm-roberta-base' #'bert-base-cased'
        self.bert_config = AutoConfig.from_pretrained(self.model_name, output_hidden_states=True)
        self.bert_tokenizer = AutoTokenizer.from_pretrained(self.model_name)
        self.bert_model = AutoModel.from_pretrained(self.model_name, config=self.bert_config)
        

        # Hyperparameters class
        @dataclass
        class HParams:
            label_vocabulary = self.vocab_label
            vocab_size_words: int = len(self.vocab_words)
            lstm_hidden_dim: int = 300
            embedding_dim_words: int = 300
            embedding_dim_lemmas: int = 300
            embedding_dim_relations: int = 300
            embedding_dim_predicates: int = 400
            embedding_dim_pos: int = 300
            gcn_output_dim: int = 143
            gcn_dropout_probability: float = 0.5
            gcn_hidden_dim: int = 250
            gcn_lstm_num_layers: int = 2
            bert_lstm_num_layers: int = 2
            bert_hidden_dim: int = self.bert_config.hidden_size
            num_classes: int = len(self.vocab_label)
            biaffine_lstm_num_layers: int = 2
            bidirectional: bool = True
            num_layers: int = 2
            dropout: float = 0.3
            lstm_dropout: float = 0.3
            vocab_size_pos_tags: int = len(self.vocab_pos_tags)
            vocab_size_lemmas: int = len(self.vocab_lemmas)
            vocab_size_predicates: int = len(self.vocab_predicates)
            device: str = self.device

        hyperparameters: HParams = HParams()

        self.net_configuration = net_configuration

        model: SRL_final_MODEL = SRL_final_MODEL(hparams=hyperparameters, configurator=net_configuration).to(self.device)
        
        # Load the SRL model
        model.load_state_dict(torch.load('srl_model/Italian/model_stored.pth', map_location=self.device))
        
        self.model: SRL_final_MODEL = model
        self.model.eval()  # set model in eval settings


    def extract_sentence_features(self, utterance):

        '''

        By using stanza's library we extract the following sentence features: words, lemmas, pos tags and
        dependency heads required to predict the roles of the sentence

        Args: 
            utterance (str): the english sentence 

        Returns:
            words (list[str]): list of words in the sentence
            lemmas (list[str]): list of words0 lemmas in the sentence
            pos_tags (list[str]): list of words' positional tags
            dep_heads (list[str]): list of words' dependencies

        '''
         
        doc = self.nlp(utterance)
        words = []
        lemmas = []
        pos_tags = []
        dep_heads = []
        features = []

        for sent in doc.sentences:
            for word in sent.words:
                words.append(word.text)
                lemmas.append(word.lemma)
                pos_tags.append(word.xpos)
                dep_heads.append(word.head)
                features.append(word.feats)

        print(words, len(words))
        print(lemmas, len(lemmas))
        print(pos_tags, len(pos_tags))
        print(dep_heads, len(dep_heads))
        print(features, len(features))

        return words, lemmas, pos_tags, dep_heads, features
    

    def get_knowledge(self, lemmas):

        '''
        The function accesses file containing the association lemma-frame that represents the 
        robot's knowledge 

        Args: -

        Return:
            dict (dict): Contains lemma-frame association
        
        '''

        verb_frame = []
        val = []
        dict = {}

        with open("./data/Frames_it.txt", 'r') as f_txt: 
            for line_no, line in enumerate(f_txt):

                # Remove the newline character
                frames = line.replace("\n","")

                # Split the line to obtain verb and associated frame
                verb_frame = frames.split(",")

                # Extract the verb
                key = verb_frame[0]

                # Extract the frame
                value = verb_frame[1]

                # --------
                # ITALIAN
                # --------
                #check if the word is in the next position of the verb
                if key == "spostare" and "grad" in lemmas:
                    value = "TURN"
                elif key == "spostare" and ("scatol" or "contenitor") in lemmas:
                    value = "CHANGE"
                elif key == "prendere" and "precauzioni" in lemmas:
                    if "precauzioni" in lemmas[lemmas.index(key)+1]:
                        value = "PAY-ATTENTION"
                elif key == "trovare" and "ti" in lemmas:
                    if "ti" in lemmas[lemmas.index(key)-1]:
                        value = "PLACE"
                elif key == "guardare" and "mi" in lemmas:
                    value = "TURN"
                elif key == "guardare" and "indietro" in lemmas:
                    value = "TURN"
                


                # --------
                # ENGLISH
                # --------
                if key == "take" and "precaution" in lemmas:
                    if "precaution" in lemmas[lemmas.index(key)+1]:
                        value = "PAY-ATTENTION"
                elif key == "pay" and "attention" in lemmas:
                    if "attention" in lemmas[lemmas.index(key)+1]:
                        value = "PAY-ATTENTION"
                elif key == "get" and ("out" or "to") in lemmas:
                    if ("out" or "to") in lemmas[lemmas.index(key)+1]:
                        value = "REACH"
                elif key == "get" and "close" in lemmas:
                    if "close" in lemmas[lemmas.index(key)+1]:
                        value = "GO"


                if key in dict.keys():
                    val = dict.get(key)
                    val.append(str(value))
                    dict[key] = val
                else:
                    val = []
                    val.append(str(value))
                    dict[key] = val
        f_txt.close()

        return dict


    def get_predicates(self, lemmas, pos_tag_l, lemma_l):

        '''
        The function accesses file containing the association lemma-frame that represents the 
        robot's knowledge 

        Args: 
            pos_tag_l (list[str]): Contains the positional tags associated to each word in the sentence
            lemma_l (list[str]): Contains the lemmas associated to each word in the sentence

        Return:
            predicates (list[str]): Contains the predicates list as expected by the SRL predictor
        
        '''
    


        # Obtain the lemma-predicate association
        dict_lemma_pred = self.get_knowledge(lemmas)

        # Initialise the predicates list
        predicates = ['_']*len(lemma_l)

        # Pass over all the elements in the pos tag list
        for idx, pos in enumerate(pos_tag_l):

            # -------------------------------
            # ---------- ITALIAN ------------
            # -------------------------------

            if pos == "VA" and pos_tag_l[idx+1] != "V":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]

            elif pos == "V":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]

            # -------------------------------
            # ---------- ENGLISH ------------
            # -------------------------------

            # Check if in the pos tag list we encounter a verb
            elif pos == "VB" and pos_tag_l[idx+1] != "VBN":

                # Take the predicate associated to the lemma of the verb
                pred = dict_lemma_pred.get(lemma_l[idx])

                if pred == None:
                    print("No predicates found")
                else:

                    # Fill the predicates list
                    predicates[idx] = pred[0]
                
            elif pos == "VBZ" and pos_tag_l[idx+1] != "VBG":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]
            
            elif pos == "VBP" and pos_tag_l[idx+1] != "VBN":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]

            elif pos == "VBG":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]

            elif pos == "VBN":
                pred = dict_lemma_pred.get(lemma_l[idx])
                if pred == None:
                    print("No predicates found")
                else:
                    predicates[idx] = pred[0]
            
        return predicates



    def predict(self, sentence):
        sent = {0: sentence}
        dataset_test: SRL_Dataset = SRL_Dataset(sentences=sent, labels=None, device=self.device, max_len=MAX_LEN, configurator=self.net_configuration, bert_model=self.bert_model, bert_tokenizer=self.bert_tokenizer)
        dataset_test.build_sample(self.vocab_words, self.vocab_pos_tags, self.vocab_lemmas, self.vocab_predicates, self.vocab_label) 
        batch_size: int = BATCH_SIZE
        dataloader_test: DataLoader = DataLoader(dataset_test, batch_size=batch_size)
        return print_output(self.model, dataloader_test, self.vocab_label)[0]
    



'''

SPEECH UNDERSTANDING

'''

class SpeechUnderstanding:

    def __init__(self, personID):

        # Load the speech act classification model
        self.model = TextPredictor.load('./speech-act-classifier')

        self.sentence_pred = SRLPrediction(DEVICE)

        self.punct_model = PunctuationModel()

        # TODO: analyse the spectogram
        self.spectogram_model = ImagePredictor.load('./speech_spectrum_classifier/image_predictor.ag')
        
        # User's ID
        self.personID = personID

        # 
        self.c = 0

        # User's language
        self.lang = ""
        
        # Textual sentence
        self.sentence = ""      
        
        # Speech act assigned to the utterance
        self.s_act = ""  

        # Create ROS message
        self.utterance_msg = Utterance()

        # Subscriber
        rospy.Subscriber('/recognition_result', Utterance, self.rec_cb, queue_size=10) 

        # Publisher
        self.decision_manager = rospy.Publisher('/decision_manager_topic_speech', FrameGroupMessage, queue_size=10)


    def rec_cb(self, rec_msg):

        '''

        Recognition callback function

        '''
        self.sent_id = rec_msg.sent_id
        self.lang = rec_msg.language
        self.sentence = rec_msg.sentence
        self.s_act = rec_msg.speech_act
        self.f_roles = rec_msg.frame_roles
        self.f_words = rec_msg.frame_words




    def restore_punctuation(self, text):

        '''

        Function that introduces punctuation in the sentence

        Args: 
            text (str): Textual utterance

        Return:
            out_sentences (list): list with longest sentences split based on the punctuation
        
        '''

        # Restore punctuation in the sentence
        restored = self.punct_model.restore_punctuation(text)

        out_sentences = []

        if ". " in restored:
            out_sentences = restored.split(". ")

        # It allows to split longest sentences
        elif ", " in restored:
            out_sentences = restored.split(", ")
        
        return out_sentences
    


    def get_user_folder(self):

        '''

        Get user's excel file to save the vocal utterances

        Args: -

        Returns:
            excel (path): The path to the user's excel file

        '''

        # Parent Directory path
        parent_dir = "./User Interactions/"
        directory = "Person_" + self.personID
        excel_file = "TEST_"+ self.personID +".xlsx"
        path_users = os.path.join(parent_dir, directory)
        excel = os.path.join(path_users, excel_file)

        return excel


    def get_frame_definition(self, words, lemmas, pos_tags, predicates, roles):

        '''
        The function prepares both frame-roles and frame-words definition in two different lists 
        to be passed to DEM module 

        Args: 
            words (list[str]): Contains the words of the sentence
            pos_tag (list[str]): Contains the positional tags associated to each word in the sentence
            lemmas (list[str]): Contains the lemmas associated to each word in the sentence
            predicates (list[str]): Contains the verbs identified in the sentence
            roles (list[str]): Contains the roles associated to each verb in the sentence

        Return:
            The output in both cases will be in the form: FRAME(arg1, arg2, ....)

            list_frame_roles (list[str]): Contains the general frame-roles definition
            list_frame_words (list[str]): Contains the frame-words definition associated to the sentence
        
        '''

        list_frame_roles = []
        list_frame_words = []

        final_frame_roles = []
        final_frame_words = []
        
        # Predicates
        for i, pred in enumerate(predicates):
            
            if pred != "_":
                list_frame_roles.append(pred) 
                list_frame_words.append(pred) 

                # Roles
                list_roles = roles['roles'].get(i)
                for idx, r in enumerate(list_roles):
                    if r == "_" and pos_tags[idx] == "PRP" and lemmas[idx] == "I":
                        list_frame_words.append(words[idx])
                    if r != "_" and r != "Goal" and r != "Purpose" and r not in list_frame_roles:
                        list_frame_roles.append(r)

                        if pos_tags[idx] != "NNS":
                            list_frame_words.append(words[idx])

                        else: 
                            list_frame_words.append(lemmas[idx])
                    
            if list_frame_roles != [] and list_frame_words != []:
                #print("YESS")
                #print('**************')
                #print(list_frame_roles)
                #print(list_frame_words)


                final_frame_roles.append(list_frame_roles)
                final_frame_words.append(list_frame_words)

                #print(frame_ro_msg)
                #print(frame_ro_msg)


                list_frame_roles = []
                list_frame_words = []

        return final_frame_roles, final_frame_words



    def predict_speech_act_SRL(self):

        '''

        Function that predicts the sentence speech act and semantic role labeling
        
        '''

        excel = self.get_user_folder()
        wb = openpyxl.Workbook()
        sheet = wb.active

        # While the node is not shutted down
        while not rospy.is_shutdown():

            #time_start = datetime.now()

            # ----------------------------
            # Starting the analysis of the speech act
            # ----------------------------

            # Check if no speech act has been assigned to the sentence
            if self.s_act == "" and self.sentence != "":

                # Apply punctuation restoration to split longest/complex sentences in case there are any
                self.complex_utt = self.restore_punctuation(self.sentence)

                # Complex sentences
                if len(self.complex_utt) > 1:
                    for i in range(len(self.complex_utt)):

                        # Speech act prediction of the sentence
                        cat = self.model.predict({'Sentence': self.complex_utt})[i]
                        if i == 0:
                            self.s_act += cat
                        else:
                            self.s_act += " + " + cat

                    #print('"Sentence":', self.complex_utt, '"Predicted Speech Act":', self.s_act)

                    #print('Time required is:')
                    #print(time.process_time() - start_SpeechAct)

                    # Fill the excel file with language, sentence and speech act information
                    self.c += 1
                    sheet.cell(row=self.c, column=1).value = self.lang
                    sheet.cell(row=self.c, column=2).value = self.sentence
                    sheet.cell(row=self.c, column=3).value = self.s_act

                # Short sentences
                else:

                    # Speech act prediction of the sentence
                    self.s_act = self.model.predict({'Sentence': [self.sentence]})[0]

                    #print('"Sentence":', self.sentence, '"Predicted Speech Act":', self.s_act)

                    #time_speech_act = time.process_time() - start_SpeechAct
                    #print('Time required for speech act prediction is:')
                    #print(time_speech_act)

                    # Fill the excel file with language, sentence and speech act information
                    self.c += 1
                    sheet.cell(row=self.c, column=1).value = self.lang
                    sheet.cell(row=self.c, column=2).value = self.sentence
                    sheet.cell(row=self.c, column=3).value = self.s_act

                    # TODO: Use the spectogram to avoid intonation issues between Information and Request 
                    # (the code needs to be adjusted)
                    if self.s_act == "Information":
                        result = self.spectogram_model.predict("./spectrogram_prova"+self.id+".jpg")
                        print(result)
                        str_result = str(result)
                        out_pred_1 = str_result.split(" ")
                        print(out_pred_1)
                        label_pred_str = out_pred_1[4]
                        out_pred_2 = label_pred_str.split("\n")
                        label_pred = out_pred_2[0]
                        print(label_pred)
                        sheet.cell(row=self.c, column=4).value = label_pred
                        if label_pred == "Request":
                            self.s_act = "Request"


                # ----------------------------
                # Starting the analysis of SRL
                # ----------------------------
                
                #sentence_pred = SRLPrediction(DEVICE)

                # If the sentence is not in english, we need to translate it to allow internal 
                # analysis by the system
                
                if self.lang != "italian" and self.lang != "english":
                    self.sentence, self.lang = translate_sentence(self.sentence, self.lang)
                    print(self.sentence)

                #time_start_SRL = time.process_time()

                # Get sentence features 
                words, lemmas, pos_tags, dep_heads, features = self.sentence_pred.extract_sentence_features(self.sentence)
                predicates = self.sentence_pred.get_predicates(lemmas, pos_tags, lemmas)
                print(features)
                roles = self.sentence_pred.predict({"words":words,
                                               "lemmas":lemmas,
                                               "pos_tags":pos_tags,
                                               "predicates":predicates,
                                               "dependency_heads":dep_heads
                                              })



                # TODO: save the predicates and roles in the excel file
                #self.c += 1
                #sheet.cell(row=self.c, column=4).append(predicates)
                #sheet.cell(row=self.c, column=5).append(roles)

                wb.save(excel)
                print("Sentence information added!")
                
                frame_roles_def, frame_words_def = self.get_frame_definition(words, lemmas, pos_tags, predicates, roles)

                #print(frame_roles_def)
                #print(frame_words_def)

                frames: list[FrameMessage] = []

                for i, data in enumerate(frame_words_def):

                    frame = FrameMessage()
                    frame.frame_roles = frame_roles_def[i] 
                    frame.frame_words = frame_words_def[i]
                    frame.full_sentence = self.sentence
                    frame.speech_act = self.s_act
                    frame.language = self.lang
                    frames.append(frame)

                frame_group_message = FrameGroupMessage()
                frame_group_message.frames = frames

                self.decision_manager.publish(frame_group_message)

                #######

                '''self.utterance_msg.sent_id = self.sent_id
                self.utterance_msg.language = self.lang
                self.utterance_msg.sentence = self.sentence
                self.utterance_msg.speech_act = self.s_act
                self.utterance_msg.frame_roles = frame_roles_def
                self.utterance_msg.frame_words = frame_words_def
                '''

                print(self.lang + "; " + self.sentence + "; " + self.s_act + "; " + str(predicates) + "; " + str(roles))
                print("Frame-Roles definition: ", frame_roles_def)
                print("Frame-Words definition: ", frame_words_def)
                
                self.sentence = ""

                '''print("***********************")
                print('Time required for Speech Act + SRL predictions is:')
                print(datetime.now() - time_start)
                print("***********************")'''
                
                




if __name__ == '__main__':
    rospy.init_node('speech_und', anonymous=True)

    # Get the ID of the user
    personID = get_user_id()

    # Create an instance of Speech Understanding
    voice_underst = SpeechUnderstanding(personID)

    # Predict both speech act and SRL
    voice_underst.predict_speech_act_SRL()

    